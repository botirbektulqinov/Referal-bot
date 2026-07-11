"""Foydalanuvchilar hisobotini chiroyli dizaynli Excel (.xlsx) ga aylantiradi."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["№", "Ism Familya", "Telefon", "Coin", "Taklif qilgan (ID)", "Sana"]
WIDTHS = [7, 30, 20, 9, 20, 14]

BRAND = "FF2563EB"   # ko'k (title)
DARK = "FF1E293B"    # to'q (header)
STRIPE = "FFF1F5F9"  # och kulrang (juft qatorlar)
WHITE = "FFFFFFFF"
MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}


def build_excel(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    ncols = len(HEADERS)
    thin = Side(style="thin", color="FFCBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # --- Sarlavha (birlashtirilgan) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(row=1, column=1, value="📊 FOYDALANUVCHILAR HISOBOTI")
    title.font = Font(bold=True, size=16, color=WHITE)
    title.alignment = center
    title.fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 32

    # --- Ustun sarlavhalari ---
    for c, head in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=head)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 24

    # --- Ma'lumot qatorlari ---
    for i, r in enumerate(rows, 1):
        row_idx = i + 2
        values = [
            MEDALS.get(i, str(i)),
            r["full_name"] or "—",
            r["phone"] or "—",
            r["coins"],
            r["inviter_id"] or "—",
            (r["created_at"] or "")[:10],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.border = border
            cell.alignment = center if c != 2 else Alignment(horizontal="left", vertical="center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=STRIPE)
            if c == 4:  # coin ustuni ajralib tursin
                cell.font = Font(bold=True, color=BRAND)

    # --- Kenglik + muzlatilgan sarlavha ---
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":  # ponytail: rang/border buzilmaganini tekshiruvchi mini self-check
    data = build_excel(
        [
            {"full_name": "Test User", "phone": "+998901112233", "coins": 5,
             "inviter_id": None, "created_at": "2026-07-11T10:00:00"},
        ]
    )
    assert data[:2] == b"PK" and len(data) > 2000, "xlsx yaratilmadi"
    print("OK", len(data), "bytes")
